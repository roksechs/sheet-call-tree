"""Extract formula cells from an openpyxl workbook."""
from __future__ import annotations

import logging
import re
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

from .formula_parser import parse_formula
from .labeler import build_label_map
from .models import CellNode, FunctionNode, NamedRefNode, RangeNode, TableRefNode

log = logging.getLogger(__name__)


def extract_formula_cells(
    path: str | Path,
) -> tuple[dict[str, FunctionNode], dict[str, object], dict[str, dict[str, object]]]:
    """Load an xlsx/xlsm file and return formula ASTs, cached values, and labels.

    Loads the workbook twice: once for formulas (data_only=False, full load) and
    once for cached values (data_only=True, read_only streaming to save memory).
    Both are iterated together in a single pass via zip().  After each stage the
    workbook objects are closed so their memory is freed before the next stage.

    Only formula cells (values starting with '=') are included as keys in
    the first dict. Constant cells appear only as RefNode values inside
    formula ASTs.

    Args:
        path: Path to the .xlsx/.xlsm file.

    Returns:
        Tuple of (formula_cells, data_values, label_map) where formula_cells maps
        'SheetName!CellRef' strings to FunctionNode AST roots, data_values maps
        cell refs to their cached scalar values, and label_map maps cell refs to
        dicts with 'row' and 'column' label values.
    """
    wb = openpyxl.load_workbook(path, data_only=False)
    wb_data = openpyxl.load_workbook(path, data_only=True, read_only=True)
    cells, data_values = _extract_cells_and_values(wb, wb_data)
    wb_data.close()  # streaming workbook no longer needed

    table_ranges = _build_table_ranges(wb)
    named_ranges = _build_named_ranges(wb)
    merged_map = _build_merged_cell_map(wb)
    bold_cells = _build_bold_cells(wb)
    wb.close()  # formula workbook no longer needed

    _populate_ref_values(cells, data_values, table_ranges, named_ranges)
    label_map = build_label_map(cells, data_values, merged_map, bold_cells)
    return cells, data_values, label_map


def extract_formula_cells_from_workbook(wb: openpyxl.Workbook) -> dict[str, FunctionNode]:
    """Extract formula cells from an already-loaded workbook.

    CellNode.expression and CellNode.outputs are left as None; call
    _populate_ref_values() after loading a data_only workbook if needed.
    """
    result: dict[str, FunctionNode] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if not isinstance(val, str) or not val.startswith("="):
                    continue
                col_letter = cell.column_letter
                row_num = cell.row
                cell_ref = f"{sheet_name}!{col_letter}{row_num}"
                ast = parse_formula(val, default_sheet=sheet_name)
                if ast is not None:
                    result[cell_ref] = ast
                else:
                    log.warning("Could not parse formula at %s: %r", cell_ref, val)

    return result


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _extract_cells_and_values(
    wb: openpyxl.Workbook,
    wb_data: openpyxl.Workbook,
) -> tuple[dict[str, FunctionNode], dict[str, object]]:
    """Extract formula cells and cell values in a single pass using zip()."""
    formula_cells: dict[str, FunctionNode] = {}
    data_values: dict[str, object] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws_data = wb_data[sheet_name]
        for row, row_data in zip(ws.iter_rows(), ws_data.iter_rows()):
            for cell, cell_data in zip(row, row_data):
                if cell_data.value is not None:
                    ref = f"{sheet_name}!{cell_data.column_letter}{cell_data.row}"
                    data_values[ref] = cell_data.value
                val = cell.value
                if isinstance(val, str) and val.startswith("="):
                    cell_ref = f"{sheet_name}!{cell.column_letter}{cell.row}"
                    ast = parse_formula(val, default_sheet=sheet_name)
                    if ast is not None:
                        formula_cells[cell_ref] = ast
                    else:
                        log.warning("Could not parse formula at %s: %r", cell_ref, val)
    return formula_cells, data_values


def _extract_all_values(wb: openpyxl.Workbook) -> dict[str, object]:
    """Return a dict of all non-None cell values from a data_only workbook."""
    result: dict[str, object] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    ref = f"{sheet_name}!{cell.column_letter}{cell.row}"
                    result[ref] = cell.value
    return result


def _build_table_ranges(wb) -> dict[str, dict]:
    """Build a map of table name → column ranges for all tables in the workbook.

    Returns: { "Table1": {"_sheet": "Sheet1", "_range": "Sheet1!A1:B10",
                          "columns": {"ColumnName": "Sheet1!B2:B10"}} }
    """
    result: dict[str, dict] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for table in ws.tables.values():
            table_ref = table.ref  # e.g. "A1:D10"
            start, end = table_ref.split(":")
            start_col = "".join(c for c in start if c.isalpha())
            start_row = int("".join(c for c in start if c.isdigit()))
            end_row = int("".join(c for c in end if c.isdigit()))
            data_start_row = start_row + 1  # first row is header
            start_col_idx = column_index_from_string(start_col)
            columns: dict[str, str] = {}
            for i, col in enumerate(table.tableColumns):
                col_letter = get_column_letter(start_col_idx + i)
                columns[col.name] = (
                    f"{sheet_name}!{col_letter}{data_start_row}:{col_letter}{end_row}"
                )
            result[table.displayName] = {
                "_sheet": sheet_name,
                "_range": f"{sheet_name}!{table_ref}",
                "columns": columns,
            }
    return result


def _build_named_ranges(wb) -> dict[str, str]:
    """Build a map of defined name → range text for all named ranges in the workbook.

    Returns: { "SalesTotal": "Sheet1!$B$2" }
    """
    result: dict[str, str] = {}
    for name in wb.defined_names:
        dn = wb.defined_names[name]
        result[name] = dn.attr_text
    return result


def _populate_ref_values(
    cells: dict[str, FunctionNode],
    data_values: dict[str, object],
    table_ranges: dict[str, dict] | None = None,
    named_ranges: dict[str, str] | None = None,
) -> None:
    """Walk every AST and fill ref node values.

    - Formula-cell refs: expression = FunctionNode of that cell; outputs from data_only.
    - Constant-cell refs: outputs = scalar from data_only; expression stays None.
    - RangeNode: cells populated from data_values for all cells in the range.
    - TableRefNode: resolved_range from table_ranges map.
    - NamedRefNode: resolved_range from named_ranges map; cell with expression/outputs if single cell.
    - Unknown refs: fields remain None.
    """
    known = set(cells)
    for fn in cells.values():
        _fill_node(fn, cells, data_values, known, table_ranges or {}, named_ranges or {})


def _fill_node(node, cells, data_values, known, table_ranges, named_ranges):
    if isinstance(node, FunctionNode):
        for arg in node.inputs:
            _fill_node(arg, cells, data_values, known, table_ranges, named_ranges)
    elif isinstance(node, CellNode):
        ref = node.cell
        if ref in known:
            node.expression = cells[ref]
            node.outputs = data_values.get(ref)
        elif ref in data_values:
            node.outputs = data_values[ref]
        # else: both fields stay None
    elif isinstance(node, RangeNode):
        node.cells = _resolve_range_cells(node.start, node.end, data_values)
    elif isinstance(node, TableRefNode):
        tbl = table_ranges.get(node.table_name)
        if tbl:
            if node.column and node.column in tbl["columns"]:
                node.resolved_range = tbl["columns"][node.column]
            elif not node.column:
                node.resolved_range = tbl["_range"]
    elif isinstance(node, NamedRefNode):
        attr_text = named_ranges.get(node.name)
        if attr_text:
            node.resolved_range = attr_text
            normalized = attr_text.replace("$", "")
            if normalized in known:
                node.cell = CellNode(
                    cell=normalized,
                    expression=cells[normalized],
                    outputs=data_values.get(normalized),
                )
            elif normalized in data_values:
                node.cell = CellNode(
                    cell=normalized,
                    outputs=data_values[normalized],
                )


def _build_merged_cell_map(wb: openpyxl.Workbook) -> dict[str, str]:
    """Map merged cell refs to top-left cell ref.

    For each merged range, every cell in the range maps to the top-left cell.
    """
    result: dict[str, str] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for merged_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left = f"{sheet_name}!{get_column_letter(min_col)}{min_row}"
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ref = f"{sheet_name}!{get_column_letter(col)}{row}"
                    if ref != top_left:
                        result[ref] = top_left
    return result


def _build_bold_cells(wb: openpyxl.Workbook) -> set[str]:
    """Return set of cell refs with bold font."""
    result: set[str] = set()
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.font and cell.font.bold:
                    result.add(f"{sheet_name}!{cell.column_letter}{cell.row}")
    return result


_CELL_REF_RE = re.compile(r"^([A-Za-z]+)(\d+)$")


def _resolve_range_cells(start: str, end: str, data_values: dict[str, object]) -> list[CellNode] | None:
    """Enumerate all cells in a rectangular range and return CellNode objects.

    Returns None if the range endpoints can't be parsed as cell references
    (e.g. whole-column ranges like A:B).
    """
    # Parse sheet and cell parts
    if "!" in start:
        sheet, start_cell = start.rsplit("!", 1)
    else:
        return None
    if "!" in end:
        _, end_cell = end.rsplit("!", 1)
    else:
        end_cell = end

    m_start = _CELL_REF_RE.match(start_cell)
    m_end = _CELL_REF_RE.match(end_cell)
    if not m_start or not m_end:
        return None

    start_col_idx = column_index_from_string(m_start.group(1))
    start_row = int(m_start.group(2))
    end_col_idx = column_index_from_string(m_end.group(1))
    end_row = int(m_end.group(2))

    result = []
    for row in range(start_row, end_row + 1):
        for col_idx in range(start_col_idx, end_col_idx + 1):
            col_letter = get_column_letter(col_idx)
            ref = f"{sheet}!{col_letter}{row}"
            result.append(CellNode(cell=ref, outputs=data_values.get(ref)))
    return result
