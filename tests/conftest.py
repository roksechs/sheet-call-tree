"""Shared fixtures: synthetic openpyxl workbooks for testing."""
import io

import openpyxl
import pytest
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableColumn


@pytest.fixture
def simple_workbook():
    """Workbook with basic formula cells used in the plan's example."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 10
    ws["A2"] = 20
    ws["C5"] = "=SUM(A1:A2)"
    ws["B10"] = "=C5+1.1"
    ws["B11"] = "=C5*2.0"
    return wb


@pytest.fixture
def simple_workbook_path(simple_workbook, tmp_path):
    """Save simple_workbook to a temp file and return the path."""
    p = tmp_path / "test.xlsx"
    simple_workbook.save(p)
    return p


@pytest.fixture
def simple_workbook_xlsm_path(simple_workbook, tmp_path):
    """Save simple_workbook as .xlsm and return the path."""
    p = tmp_path / "test.xlsm"
    simple_workbook.save(p)
    return p


@pytest.fixture
def multi_sheet_workbook():
    """Workbook with cross-sheet formula references."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = 5
    ws1["B1"] = "=Sheet2!A1 + A1"

    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "=Sheet1!A1 * 2"
    return wb


@pytest.fixture
def multi_sheet_workbook_path(multi_sheet_workbook, tmp_path):
    p = tmp_path / "multi.xlsx"
    multi_sheet_workbook.save(p)
    return p


@pytest.fixture
def circular_workbook():
    """Workbook containing a circular reference A1 → B1 → A1."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "=B1+1"
    ws["B1"] = "=A1+1"
    return wb


@pytest.fixture
def circular_workbook_path(circular_workbook, tmp_path):
    p = tmp_path / "circular.xlsx"
    circular_workbook.save(p)
    return p


@pytest.fixture
def table_workbook():
    """Workbook with an Excel table (Table1) containing Name and Amount columns."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Name"
    ws["B1"] = "Amount"
    ws["A2"] = "Alice"
    ws["B2"] = 100
    ws["A3"] = "Bob"
    ws["B3"] = 200
    # Formula using a table structured reference
    ws["C2"] = "=SUM(Table1[Amount])"
    # Formula using this-row reference
    ws["D2"] = "=Table1[@Amount]*2"
    table = Table(displayName="Table1", ref="A1:B3")
    table.tableColumns = [TableColumn(id=1, name="Name"), TableColumn(id=2, name="Amount")]
    ws.add_table(table)
    return wb


@pytest.fixture
def table_workbook_path(table_workbook, tmp_path):
    p = tmp_path / "table.xlsx"
    table_workbook.save(p)
    return p


@pytest.fixture
def labeled_workbook():
    """Workbook with clear label structure for testing label detection."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # Header row (text, high entropy vs data rows)
    ws["A1"] = "Name"
    ws["B1"] = "Q1"
    ws["C1"] = "Q2"
    ws["D1"] = "Total"
    # Data rows (formulas, low entropy)
    ws["A2"] = "Alice"
    ws["B2"] = 100
    ws["C2"] = 200
    ws["D2"] = "=SUM(B2:C2)"
    ws["A3"] = "Bob"
    ws["B3"] = 150
    ws["C3"] = 250
    ws["D3"] = "=SUM(B3:C3)"
    # Summary row (different formula pattern)
    ws["A4"] = "Total"
    ws["B4"] = "=SUM(B2:B3)"
    ws["C4"] = "=SUM(C2:C3)"
    ws["D4"] = "=SUM(D2:D3)"
    return wb


@pytest.fixture
def labeled_workbook_path(labeled_workbook, tmp_path):
    p = tmp_path / "labeled.xlsx"
    labeled_workbook.save(p)
    return p


@pytest.fixture
def merged_label_workbook():
    """Workbook with merged header cells and enough data for classifier."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Revenue"
    ws.merge_cells("A1:D1")
    ws["A2"] = "Region"
    ws["B2"] = "Q1"
    ws["C2"] = "Q2"
    ws["D2"] = "Q3"
    ws["A3"] = "North"
    ws["B3"] = 100
    ws["C3"] = 200
    ws["D3"] = "=SUM(B3:C3)"
    ws["A4"] = "South"
    ws["B4"] = 150
    ws["C4"] = 250
    ws["D4"] = "=SUM(B4:C4)"
    ws["A5"] = "Total"
    ws["B5"] = "=SUM(B3:B4)"
    ws["C5"] = "=SUM(C3:C4)"
    ws["D5"] = "=SUM(D3:D4)"
    return wb


@pytest.fixture
def merged_label_workbook_path(merged_label_workbook, tmp_path):
    p = tmp_path / "merged.xlsx"
    merged_label_workbook.save(p)
    return p


@pytest.fixture
def named_range_workbook():
    """Workbook with a named range 'SalesTotal' pointing to Sheet1!B2."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["B2"] = 500
    ws["C2"] = "=SalesTotal+10"
    dn = DefinedName(name="SalesTotal", attr_text="Sheet1!$B$2")
    wb.defined_names["SalesTotal"] = dn
    return wb


@pytest.fixture
def named_range_workbook_path(named_range_workbook, tmp_path):
    p = tmp_path / "named.xlsx"
    named_range_workbook.save(p)
    return p
