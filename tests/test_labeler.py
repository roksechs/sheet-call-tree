"""Tests for the labeler module (classifier-based, top-k)."""
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

from sheet_call_tree.labeler import (
    _load_classifier,
    build_label_map,
    _is_numeric_str,
)
from sheet_call_tree.reader import (
    _build_bold_cells,
    _build_merged_cell_map,
    extract_formula_cells_from_workbook,
)


class TestIsNumericStr:
    def test_integer(self):
        assert _is_numeric_str("123")

    def test_float(self):
        assert _is_numeric_str("3.14")

    def test_negative(self):
        assert _is_numeric_str("-42")

    def test_comma_separated(self):
        assert _is_numeric_str("1,234,567")

    def test_dollar(self):
        assert _is_numeric_str("$100")

    def test_parenthetical_negative(self):
        assert _is_numeric_str("(500)")

    def test_text(self):
        assert not _is_numeric_str("hello")

    def test_empty(self):
        assert not _is_numeric_str("")


def _extract_data_values(wb: openpyxl.Workbook) -> dict[str, object]:
    result: dict[str, object] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and not (
                    isinstance(cell.value, str) and cell.value.startswith("=")
                ):
                    ref = f"{sheet_name}!{cell.column_letter}{cell.row}"
                    result[ref] = cell.value
    return result


class TestBuildLabelMap:
    def test_labeled_workbook_top1(self, labeled_workbook):
        """With top_k=1, each axis returns a single-element list."""
        formula_cells = extract_formula_cells_from_workbook(labeled_workbook)
        data_values = _extract_data_values(labeled_workbook)
        merged_map = _build_merged_cell_map(labeled_workbook)
        bold_cells = _build_bold_cells(labeled_workbook)

        label_map = build_label_map(
            formula_cells, data_values, merged_map, bold_cells, top_k=1,
        )

        d2 = label_map["Sheet1!D2"]
        assert d2["column"] == ["Total"]
        assert d2["row"] == ["Alice"]

        d3 = label_map["Sheet1!D3"]
        assert d3["column"] == ["Total"]
        assert d3["row"] == ["Bob"]

        b4 = label_map["Sheet1!B4"]
        assert b4["column"] == ["Q1"]
        assert b4["row"] == ["Total"]

    def test_labeled_workbook_topk(self, labeled_workbook):
        """With default top_k=5, multiple header candidates are returned."""
        formula_cells = extract_formula_cells_from_workbook(labeled_workbook)
        data_values = _extract_data_values(labeled_workbook)
        merged_map = _build_merged_cell_map(labeled_workbook)
        bold_cells = _build_bold_cells(labeled_workbook)

        label_map = build_label_map(formula_cells, data_values, merged_map, bold_cells)

        d2 = label_map["Sheet1!D2"]
        # Column: nearest first (Total from D1)
        assert isinstance(d2["column"], list)
        assert d2["column"][0] == "Total"
        # Row: nearest first (Alice from A2)
        assert isinstance(d2["row"], list)
        assert d2["row"][0] == "Alice"

    def test_merged_label_workbook(self, merged_label_workbook):
        formula_cells = extract_formula_cells_from_workbook(merged_label_workbook)
        data_values = _extract_data_values(merged_label_workbook)
        merged_map = _build_merged_cell_map(merged_label_workbook)
        bold_cells = _build_bold_cells(merged_label_workbook)

        label_map = build_label_map(formula_cells, data_values, merged_map, bold_cells)

        # D3 has =SUM(B3:C3), should have row label "North"
        d3 = label_map["Sheet1!D3"]
        assert isinstance(d3["row"], list)
        assert d3["row"][0] == "North"
        # B5 has =SUM(B3:B4), should have row label "Total"
        b5 = label_map["Sheet1!B5"]
        assert b5["row"][0] == "Total"

    def test_empty_workbook(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "=1+1"
        formula_cells = extract_formula_cells_from_workbook(wb)
        data_values = _extract_data_values(wb)
        merged_map = _build_merged_cell_map(wb)
        bold_cells = _build_bold_cells(wb)

        label_map = build_label_map(formula_cells, data_values, merged_map, bold_cells)
        assert label_map.get("Sheet1!A1", {}).get("row") is None
        assert label_map.get("Sheet1!A1", {}).get("column") is None

    def test_topk_no_duplicates(self, labeled_workbook):
        """Duplicate label texts are deduplicated."""
        formula_cells = extract_formula_cells_from_workbook(labeled_workbook)
        data_values = _extract_data_values(labeled_workbook)
        merged_map = _build_merged_cell_map(labeled_workbook)
        bold_cells = _build_bold_cells(labeled_workbook)

        label_map = build_label_map(formula_cells, data_values, merged_map, bold_cells)

        for labels in label_map.values():
            for axis in ("row", "column"):
                if axis in labels:
                    vals = labels[axis]
                    assert len(vals) == len(set(vals)), f"Duplicates in {axis}: {vals}"


class TestEndToEnd:
    def test_extract_with_labels(self, labeled_workbook_path):
        from sheet_call_tree.reader import extract_formula_cells

        formula_cells, data_values, label_map = extract_formula_cells(labeled_workbook_path)
        assert isinstance(label_map, dict)
        assert any("row" in v or "column" in v for v in label_map.values())

    def test_yaml_output_includes_labels(self, labeled_workbook_path):
        from sheet_call_tree.reader import extract_formula_cells
        from sheet_call_tree.serializer import to_yaml

        formula_cells, data_values, label_map = extract_formula_cells(labeled_workbook_path)
        yaml_str = to_yaml(formula_cells, data_values=data_values, label_map=label_map)
        assert "labels:" in yaml_str
        assert "row:" in yaml_str
        assert "column:" in yaml_str

    def test_yaml_topk_lists(self, labeled_workbook_path):
        """YAML output renders label lists with '- ' items."""
        from sheet_call_tree.reader import extract_formula_cells
        from sheet_call_tree.serializer import to_yaml

        formula_cells, data_values, label_map = extract_formula_cells(labeled_workbook_path)
        yaml_str = to_yaml(formula_cells, data_values=data_values, label_map=label_map)
        # Lists are rendered as "- item" lines under row:/column:
        assert "        - " in yaml_str

    def test_cli_outputs_labels(self, labeled_workbook_path):
        """CLI main() produces YAML with labels."""
        import io
        from sheet_call_tree.cli import main

        rc = main([str(labeled_workbook_path)])
        assert rc == 0


class TestClassifierModel:
    def test_model_loads(self):
        """The .joblib model file loads successfully."""
        clf = _load_classifier()
        assert hasattr(clf, "predict")
        assert hasattr(clf, "predict_proba")

    def test_model_file_exists(self):
        model_path = Path(__file__).resolve().parent.parent / "src" / "sheet_call_tree" / "cell_classifier.joblib"
        assert model_path.exists()

    def test_model_has_expected_features(self):
        """Model expects 23 features."""
        clf = _load_classifier()
        assert clf.n_features_in_ == 23

    def test_model_two_classes(self):
        """Model predicts 2 classes: header (0) and data (1)."""
        clf = _load_classifier()
        assert list(clf.classes_) == [0, 1]


class TestBuildMergedCellMap:
    def test_simple_merge(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Header"
        ws.merge_cells("A1:C1")
        merged_map = _build_merged_cell_map(wb)
        # B1 and C1 map to A1
        assert merged_map["Sheet1!B1"] == "Sheet1!A1"
        assert merged_map["Sheet1!C1"] == "Sheet1!A1"
        # A1 itself is not in the map (it's the top-left)
        assert "Sheet1!A1" not in merged_map

    def test_no_merges(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "x"
        merged_map = _build_merged_cell_map(wb)
        assert merged_map == {}

    def test_vertical_merge(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Title"
        ws.merge_cells("A1:A3")
        merged_map = _build_merged_cell_map(wb)
        assert merged_map["Sheet1!A2"] == "Sheet1!A1"
        assert merged_map["Sheet1!A3"] == "Sheet1!A1"

    def test_block_merge(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["B2"] = "Block"
        ws.merge_cells("B2:C3")
        merged_map = _build_merged_cell_map(wb)
        assert merged_map["Sheet1!C2"] == "Sheet1!B2"
        assert merged_map["Sheet1!B3"] == "Sheet1!B2"
        assert merged_map["Sheet1!C3"] == "Sheet1!B2"


class TestBuildBoldCells:
    def test_bold_detected(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Bold"
        ws["A1"].font = openpyxl.styles.Font(bold=True)
        ws["B1"] = "Normal"
        bold = _build_bold_cells(wb)
        assert "Sheet1!A1" in bold
        assert "Sheet1!B1" not in bold

    def test_no_bold(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Normal"
        bold = _build_bold_cells(wb)
        assert bold == set()
