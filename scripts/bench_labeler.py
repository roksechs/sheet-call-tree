"""Benchmark labeler inference speed at various sheet sizes."""
from __future__ import annotations

import time

import openpyxl

from sheet_call_tree.labeler import build_label_map, _load_classifier
from sheet_call_tree.reader import (
    _build_bold_cells,
    _build_merged_cell_map,
    extract_formula_cells_from_workbook,
)


def _make_workbook(n_rows: int, n_cols: int) -> openpyxl.Workbook:
    """Create a workbook with header row + data rows with formulas."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Header row
    for c in range(1, n_cols + 1):
        ws.cell(row=1, column=c, value=f"Header_{c}")

    # Row labels in col A + data/formulas
    for r in range(2, n_rows + 1):
        ws.cell(row=r, column=1, value=f"Row_{r}")
        for c in range(2, n_cols + 1):
            if c == n_cols:
                # Last column = SUM formula
                start = openpyxl.utils.get_column_letter(2)
                end = openpyxl.utils.get_column_letter(n_cols - 1)
                ws.cell(row=r, column=c, value=f"=SUM({start}{r}:{end}{r})")
            else:
                ws.cell(row=r, column=c, value=r * c)

    return wb


def _extract_data_values(wb):
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and not (
                    isinstance(cell.value, str) and str(cell.value).startswith("=")
                ):
                    ref = f"{sheet_name}!{cell.column_letter}{cell.row}"
                    result[ref] = cell.value
    return result


def bench(n_rows: int, n_cols: int, n_runs: int = 3):
    wb = _make_workbook(n_rows, n_cols)
    total_cells = n_rows * n_cols

    formula_cells = extract_formula_cells_from_workbook(wb)
    data_values = _extract_data_values(wb)
    merged_map = _build_merged_cell_map(wb)
    bold_cells = _build_bold_cells(wb)

    # Warm up classifier loading
    _load_classifier()

    times = []
    for _ in range(n_runs):
        t0 = time.perf_counter()
        label_map = build_label_map(formula_cells, data_values, merged_map, bold_cells)
        t1 = time.perf_counter()
        times.append(t1 - t0)

    avg_ms = sum(times) / len(times) * 1000
    min_ms = min(times) * 1000
    n_labeled = len(label_map)
    print(f"  {n_rows:5d}x{n_cols:<4d} ({total_cells:7,d} cells, {len(formula_cells):5d} formulas) "
          f"→ {n_labeled:5d} labeled | avg {avg_ms:7.1f}ms  min {min_ms:7.1f}ms")


def main():
    print("Labeler benchmark (CPU, RandomForest)")
    print("=" * 75)

    configs = [
        (10, 5),       # 50 cells — tiny
        (50, 10),      # 500 cells — small
        (100, 20),     # 2,000 cells — medium
        (500, 20),     # 10,000 cells — large
        (1000, 30),    # 30,000 cells — very large
        (2000, 50),    # 100,000 cells — huge
        (5000, 50),    # 250,000 cells — extreme
    ]

    for n_rows, n_cols in configs:
        bench(n_rows, n_cols)


if __name__ == "__main__":
    main()
